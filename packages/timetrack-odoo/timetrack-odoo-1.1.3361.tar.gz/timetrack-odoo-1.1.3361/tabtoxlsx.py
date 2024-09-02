#! /usr/bin/env python3

from datetime import datetime as Time
from datetime import date as Date
import datetime
from collections import OrderedDict
"""
This script converts table-like data (list of dicts) from and to Excel .xlsx files.
It is a helper for the tabtotext.py module using openpyxl (see fallback tabxlsx.py).
Defaults to read any given file.xy rewriting it as file.xy.xlsx!
"""

__copyright__ = "(C) 2017-2024 Guido Draheim, licensed under the Apache License 2.0"""
__version__ = "1.6.3361"

import logging
from typing import TYPE_CHECKING, cast, Union, Dict, List, Any, Sequence, Iterable, Optional
from tabtotext import JSONList, JSONDict, TabSheet
from tabtotext import ColSortList, RowSortList, LegendList, RowSortCallable, ColSortCallable, unmatched
from tabtotext import FormatCSV, FormatJSONItem, FormatsDict
from tabtools import currency_default

try:
    from openpyxl import Workbook, load_workbook  # type: ignore
    from openpyxl.worksheet.worksheet import Worksheet  # type: ignore
    from openpyxl.styles.cell_style import CellStyle as Style  # type: ignore
    from openpyxl.styles.alignment import Alignment  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
except ImportError:
    from tabxlsx import Workbook, Worksheet, CellStyle as Style, Alignment, get_column_letter
    from tabxlsx import load_workbook  # type: ignore

if TYPE_CHECKING:
    from tabxlsx import Workbook as WorkbookType
else:
    WorkbookType = Workbook


SECTION = "data"
MINWIDTH = 4
MAXCOL = 1000
MAXROWS = 100000
NIX = ""

# Excel encodes empty-string as nonexistant cell.
# Since we want to encode None as empty cell (to allow numeric computations), we assign a value for empty-string.
# Note that other people hat recommended "NA()" for null which shows as "N#A" and it is correctly transferred
# into a database as NULL by Excel itself. However in your formulas you would need to skip those cells any
# numeric operation with a numeric value and some NA() returns NA()

_Empty_String = " "

logg = logging.getLogger("TABTOXLSX")

def set_cell(ws: Worksheet, row: int, col: int, value: Any, style: Style) -> None:  # type: ignore
    coordinate = {"column": col + 1, "row": row + 1}
    ws.cell(**coordinate).value = value
    # ws.cell(**coordinate).font = style.font
    # ws.cell(**coordinate).fill = style.fill
    # ws.cell(**coordinate).border = style.border
    ws.cell(**coordinate).alignment = style.alignment
    ws.cell(**coordinate).number_format = style.number_format
    ws.cell(**coordinate).protection = style.protection
def set_width(ws: Worksheet, col: int, width: int) -> None:  # type: ignore
    ws.column_dimensions[get_column_letter(col + 1)].width = width


def saveToXLSXx(filename: str, result: Union[JSONList, JSONDict], sorts: RowSortList = [],  #
                formats: Dict[str, str] = {}, legend: LegendList = [], section: str = NIX,  #
                reorder: ColSortList = []) -> None:
    if isinstance(result, Dict):
        result = [result]
    saveToXLSX(filename, result, sorts, formats, legend=legend, section=section, reorder=reorder)

def saveToXLSX(filename: str, result: JSONList,
               sorts: RowSortList = [], formats: Dict[str, str] = {}, selected: List[str] = [],
               legend: LegendList = [], section: str = NIX, reorder: ColSortList = []) -> None:
    """ old-style RowSortList and FormatsDict assembled into headers with microsyntax """
    headers: List[str] = []
    sorting: RowSortList = []
    formatter: FormatsDict = {}
    if isinstance(sorts, Sequence) and isinstance(formats, dict):
        sortheaders: List[str] = []
        for m, header in enumerate(sorts):
            cols: List[str] = []
            for headercol in header.split("|"):
                if "@" in headercol:
                    name, suffix = headercol.split("@", 1)
                    if suffix:
                        renames = "@" + suffix
                else:
                    name, renames = headercol, ""
                sortheaders += [name]
                if name in formats:
                    cols += [name + ":" + formats[name] + renames]
                else:
                    cols += [name + renames]
            headers += ["|".join(cols)]
        logg.info("headers = %s", headers)
        logg.info("sorting = %s", sortheaders)
        sorting = sortheaders
    else:
        sorting = sorts
        formatter = formats
    save_tabtoXLSX(filename, result, headers, selected, legend=legend, section=section,  # ....
                   reorder=reorder, sorts=sorting, formatter=formatter)

def tabtoXLSX(filename: str, data: Iterable[JSONDict], headers: List[str] = [], selected: List[str] = [],  # ..
              *, legend: List[str] = [], minwidth: int = 0, section: str = NIX) -> str:
    return save_tabtoXLSX(filename, data, headers, selected, legend=legend, section=section)

def save_tabtoXLSX(filename: str, data: Iterable[JSONDict], headers: List[str] = [], selected: List[str] = [],  # ..
                   *, legend: LegendList = [], minwidth: int = 0, section: str = NIX,
                   reorder: ColSortList = [], sorts: RowSortList = [], formatter: FormatsDict = {}) -> str:
    workbook = tabto_workbook(data, headers, selected,
                              legend=legend, minwidth=minwidth, section=section,
                              reorder=reorder, sorts=sorts, formatter=formatter)
    workbook.save(filename)
    return "XLSX"

def tabto_workbook(data: Iterable[JSONDict], headers: List[str] = [], selected: List[str] = [],  # ..
                   *, legend: LegendList = [], minwidth: int = 0, section: str = NIX,
                   reorder: ColSortList = [], sorts: RowSortList = [], formatter: FormatsDict = {},
                   workbook: Optional[WorkbookType] = None) -> WorkbookType:
    minwidth = minwidth or MINWIDTH
    logg.debug("tabtoXLSX:")
    renameheaders: Dict[str, str] = {}
    showheaders: List[str] = []
    sortheaders: List[str] = []
    formats: Dict[str, str] = {}
    combine: Dict[str, List[str]] = {}
    freehdrs: Dict[str, str] = {}
    for header in headers:
        combines = ""
        for selheader in header.split("|"):
            if "@" in selheader:
                selcol, rename = selheader.split("@", 1)
            else:
                selcol, rename = selheader, ""
            if "{" in selcol and "{:" not in selcol:
                names3: List[str] = []
                freeparts = selcol.split("{")
                for freepart in freeparts[1:]:
                    colon3, brace3 = freepart.find(":"), freepart.find("}")
                    if brace3 == -1:
                        logg.error("no closing '}' for '{%s' in %s", freepart, selcol)
                        continue
                    end3 = brace3 if colon3 == -1 else min(colon3, brace3)
                    name3 = freepart[:end3]
                    names3.append(name3)
                name = " ".join(names3)
                freehdrs[name] = selcol
            elif ":" in selcol:
                name, form = selcol.split(":", 1)
                if isinstance(formats, dict):
                    fmt = form if "{" in form else ("{:" + form + "}")
                    formats[name] = fmt.replace("i}", "n}").replace("u}", "n}").replace("r}", "s}").replace("a}", "s}")
            else:
                name = selcol
            showheaders += [name]  # headers make a default column order
            if rename:
                sortheaders += [name]  # headers does not sort anymore
            if not combines:
                combines = name
            elif combines not in combine:
                combine[combines] = [name]
            elif name not in combine[combines]:
                combine[combines] += [name]
            if rename:
                renameheaders[name] = rename
    logg.debug("renameheaders = %s", renameheaders)
    logg.debug("sortheaders = %s", sortheaders)
    logg.debug("formats = %s", formats)
    logg.debug("combine = %s", combine)
    logg.debug("freehdrs = %s", freehdrs)
    combined: Dict[str, List[str]] = {}
    renaming: Dict[str, str] = {}
    filtered: Dict[str, str] = {}
    selcols: List[str] = []
    freecols: Dict[str, str] = {}
    for selecheader in selected:
        combines = ""
        for selec in selecheader.split("|"):
            if "@" in selec:
                selcol, rename = selec.split("@", 1)
            else:
                selcol, rename = selec, ""
            if "{" in selcol and "{:" not in selcol:
                names4: List[str] = []
                freeparts = selcol.split("{")
                for freepart in freeparts[1:]:
                    colon4, brace4 = freepart.find(":"), freepart.find("}")
                    if brace4 == -1:
                        logg.error("no closing '}' for '{%s' in %s", freepart, selcol)
                        continue
                    end4 = brace4 if colon4 == -1 else min(colon4, brace4)
                    name4 = freepart[:end4]
                    names4.append(name4)
                name = " ".join(names4)
                freecols[name] = selcol
            elif ":" in selcol:
                name, form = selcol.split(":", 1)
                if isinstance(formats, dict):
                    fmt = form if "{" in form else ("{:" + form + "}")
                    formats[name] = fmt.replace("i}", "n}").replace("u}", "n}").replace("r}", "s}").replace("a}", "s}")
            else:
                name = selcol
            if "<" in name:
                name, cond = name.split("<", 1)
                filtered[name] = "<" + cond
            elif ">" in name:
                name, cond = name.split(">", 1)
                filtered[name] = ">" + cond
            elif "=" in name:
                name, cond = name.split("=", 1)
                filtered[name] = "=" + cond
            selcols.append(name)
            if rename:
                renaming[name] = rename
            if not combines:
                combines = name
            elif combines not in combined:
                combined[combines] = [name]
            elif combines not in combined[combines]:
                combined[combines] += [name]
    logg.debug("combined = %s", combined)
    logg.debug("renaming = %s", renaming)
    logg.debug("filtered = %s", filtered)
    logg.debug("selcols = %s", selcols)
    logg.debug("freecols = %s", freecols)
    if not selected:
        combined = combine  # argument
        freecols = freehdrs
        renaming = renameheaders
        logg.debug("combined : %s", combined)
        logg.debug("freecols : %s", freecols)
        logg.debug("renaming : %s", renaming)
    newsorts: Dict[str, str] = {}
    colnames: Dict[str, str] = {}
    for name, rename in renaming.items():
        if "@" in rename:
            newname, newsort = rename.split("@", 1)
        elif rename and rename[0].isalpha():
            newname, newsort = rename, ""
        else:
            newname, newsort = "", rename
        if newname:
            colnames[name] = newname
            if name in formats:
                formats[newname] = formats[name]
        if newsort:
            newsorts[name] = newsort
    logg.debug("newsorts = %s", newsorts)
    logg.debug("colnames = %s", colnames)
    if sorts:
        sortcolumns = sorts
    else:
        sortcolumns = [(name if name not in colnames else colnames[name]) for name in (selcols or sortheaders)]
        if newsorts:
            for num, name in enumerate(sortcolumns):
                if name not in newsorts:
                    newsorts[name] = ("@" * len(str(num)) + str(num))
            sortcolumns = sorted(newsorts, key=lambda x: newsorts[x])
            logg.debug("sortcolumns : %s", sortcolumns)
    format: FormatJSONItem
    if formatter and isinstance(formatter, FormatJSONItem):
        format = formatter
    else:
        logg.debug("formats = %s", formats)
        format = FormatCSV(formats)
    if legend:
        logg.debug("legend is ignored for CSV output")
    selcolumns = [(name if name not in colnames else colnames[name]) for name in (selcols)]
    selheaders = [(name if name not in colnames else colnames[name]) for name in (showheaders)]
    sortkey = ColSortCallable(selcolumns or sorts or selheaders, reorder)
    sortrow = RowSortCallable(sortcolumns)
    rows: List[JSONDict] = []
    cols: Dict[str, int] = {}
    for num, item in enumerate(data):
        row: JSONDict = {}
        if "#" in selcols:
            row["#"] = num + 1
            cols["#"] = len(str(num + 1))
        skip = False
        for name, value in item.items():
            selname = name
            if name in renameheaders and renameheaders[name] in selcols:
                selname = renameheaders[name]
            if selcols and selname not in selcols and "*" not in selcols:
                continue
            try:
                if name in filtered:
                    skip = skip or unmatched(value, filtered[name])
            except: pass
            colname = selname if selname not in colnames else colnames[selname]
            row[colname] = value  # do not format the value here!
            oldlen = cols[colname] if colname in cols else max(minwidth, len(colname))
            cols[colname] = max(oldlen, len(format(colname, value)))
        for freecol, freeformat in freecols.items():
            try:
                freenames = freecol.split(" ")
                freeitem: JSONDict = dict([(freename, "") for freename in freenames])
                for name, value in item.items():
                    itemname = name
                    if name in renameheaders and renameheaders[name] in freenames:
                        itemname = renameheaders[name]
                    if itemname in freenames:
                        freeitem[itemname] = format(name, value)
                value = freeformat.format(**freeitem)
                colname = freecol if freecol not in colnames else colnames[freecol]
                row[colname] = value
                oldlen = cols[colname] if colname in cols else max(minwidth, len(colname))
                cols[colname] = max(oldlen, len(value))
            except Exception as e:
                logg.info("formatting '%s' at %s bad for:\n\t%s", freeformat, e, item)
        if not skip:
            rows.append(row)
    if isinstance(legend, dict):
        newlegend = OrderedDict()
        for name in sorted(legend.keys(), key=sortkey):
            newlegend[name] = legend[name]
        legend = newlegend
    #
    sortedrows = list(sorted(rows, key=sortrow))
    sortedcols = list(sorted(cols.keys(), key=sortkey))
    return make_workbook(sortedrows, sortedcols, cols, formats, legend=legend, section=section, workbook=workbook)

def make_workbook(rows: JSONList, cols: List[str], colwidth: Dict[str, int],
                  formats: Dict[str, str], legend: LegendList, section: str = NIX,
                  workbook: Optional[WorkbookType] = None) -> WorkbookType:
    workbook = workbook or Workbook()
    ws = workbook.active
    row = 0
    ws.title = section or SECTION
    style = Style()
    hdr_style = Style()
    hdr_style.number_format = 'General'
    hdr_style.alignment = Alignment(horizontal='right')
    txt_style = Style()
    txt_style.number_format = 'General'
    txt_style.alignment = Alignment(horizontal='left')
    date_style = Style()
    date_style.number_format = 'yyyy-mm-dd'
    date_style.alignment = Alignment(horizontal='right')
    time_style = Style()
    time_style.number_format = 'yyyy-mm-dd hh:mm'
    time_style.alignment = Alignment(horizontal='right')
    num_style = Style()
    num_style.number_format = '#,##0.00'
    num_style.alignment = Alignment(horizontal='right')
    int_style = Style()
    int_style.number_format = '#,##0'
    int_style.alignment = Alignment(horizontal='right')
    eur_style = Style()
    eur_style.number_format = '#,##0.00%c' % currency_default
    eur_style.alignment = Alignment(horizontal='right')
    col = 0
    for name in cols:
        set_cell(ws, row, col, name, hdr_style)
        set_width(ws, col, colwidth[name] + 1 + int(colwidth[name] / 3))
        col += 1
    row += 1
    for item in rows:
        values: JSONDict = dict([(name, "") for name in cols])
        for name, value in item.items():
            values[name] = value
        col = 0
        for name in cols:
            value = values[name]
            if value is None:
                # <c r="A2" s="0" t="str"><f aca="false">&quot;&quot;</f><v></v></c>
                set_cell(ws, row, col, _Empty_String, txt_style)
            elif isinstance(value, Time):
                set_cell(ws, row, col, value, time_style)
            elif isinstance(value, Date):
                set_cell(ws, row, col, value, date_style)
            elif isinstance(value, int):
                set_cell(ws, row, col, value, int_style)
            elif isinstance(value, float):
                if name in formats:
                    if "$}" in formats[name]:
                        set_cell(ws, row, col, value, eur_style)
                    else:
                        set_cell(ws, row, col, value, num_style)
                else:
                    set_cell(ws, row, col, value, num_style)
            else:
                if not value:
                    set_cell(ws, row, col, _Empty_String, txt_style)
                else:
                    set_cell(ws, row, col, value, txt_style)
            col += 1
        row += 1
    if legend and False:
        ws = workbook.create_sheet()
        ws.title = "legend"
        if isinstance(legend, str):
            set_cell(ws, 0, 1, legend, txt_style)
        elif isinstance(legend, dict):
            for row, name in enumerate(legend):
                set_cell(ws, row, 0, name, txt_style)
                set_cell(ws, row, 1, legend[name], txt_style)
        else:
            for row, line in enumerate(legend):
                set_cell(ws, row, 1, line, txt_style)
    if legend and True:
        if isinstance(legend, str):
            text = '="%s"' % legend.replace('"', "'")
            set_cell(ws, row + 0, 1, text, txt_style)
        elif isinstance(legend, dict):
            for num, name in enumerate(legend):
                text = '="%s"' % legend[name].replace('"', "'")
                set_cell(ws, row + num, 0, name, txt_style)
                set_cell(ws, row + num, 1, text, txt_style)
        else:
            for num, line in enumerate(legend):
                text = '="%s"' % line.replace('"', "'")
                set_cell(ws, row + num, 0, text, txt_style)
    return cast(WorkbookType, workbook)

def readFromXLSX(filename: str, section: str = NIX) -> JSONList:
    tablist = tablistfileXLSX(filename)
    if section:
        for tabsheet in tablist:
            if tabsheet.title == section:
                return tabsheet.data
        return []
    return tablist[0].data if tablist else []
def tablistfileXLSX(filename: str) -> List[TabSheet]:
    workbook = load_workbook(filename)
    return tablist_workbook(workbook)
def tablist_workbook(workbook: Workbook, section: str = NIX) -> List[TabSheet]:  # type: ignore[no-any-unimported]
    tab: List[TabSheet] = []
    for ws in workbook.worksheets:
        title = ws.title
        cols: List[str] = []
        for col in range(MAXCOL):
            header = ws.cell(row=1, column=col + 1)
            if header.value is None:
                break
            name = header.value
            if name is None:
                break
            cols.append(str(name))
        logg.debug("xlsx found %s cols\n\t%s", len(cols), cols)
        data: List[JSONDict] = []
        for atrow in range(MAXROWS):
            record = []
            found = 0
            for atcol in range(len(cols)):
                cell = ws.cell(row=atrow + 2, column=atcol + 1)
                if cell.data_type in ["f"]:
                    continue
                value = cell.value
                # logg.debug("[%i,%si] cell.value = %s", atcol, atrow, value)
                if value is not None:
                    found += 1
                if isinstance(value, str) and value == " ":
                    value = ""
                record.append(value)
            if not found:
                break
            newrow = dict(zip(cols, record))
            data.append(newrow)  # type: ignore[arg-type]
        tab.append(TabSheet(data, cols, title))
    return tab

def tablistmake_workbook(tablist: List[TabSheet], selected: List[str] = [], minwidth: int = 0) -> Optional[WorkbookType]:
    workbook: Optional[WorkbookType] = None
    for tabsheet in tablist:
        if workbook is not None:
            ws = workbook.create_sheet()
            try:
                workbook.active = ws  # type: ignore[misc]
            except Exception as e:
                logg.warning("could not set active: %s", e)
        work = tabto_workbook(tabsheet.data, tabsheet.headers, selected,
                              minwidth=minwidth, section=tabsheet.title,
                              workbook=workbook)
        if workbook is None:
            workbook = work
    return workbook

if __name__ == "__main__":
    from tabtotext import tablistfile
    from os.path import splitext
    DONE = (logging.WARNING + logging.ERROR) // 2
    logging.addLevelName(DONE, "DONE")
    from optparse import OptionParser
    cmdline = OptionParser("%prog [help|files...]", epilog=__doc__, version=__version__)
    cmdline.formatter.max_help_position = 29
    cmdline.add_option("-v", "--verbose", action="count", default=0, help="more verbose logging")
    cmdline.add_option("-^", "--quiet", action="count", default=0, help="less verbose logging")
    cmdline.add_option("-L", "--labels", metavar="LIST", action="append",  # ..
                       help="select columns to show (a|x=b)", default=[])
    cmdline.add_option("-i", "--inputformat", metavar="FMT", help="fix input format (instead of autodetection)", default="")
    opt, args = cmdline.parse_args()
    logging.basicConfig(level=max(0, logging.WARNING - 10 * opt.verbose + 10 * opt.quiet))
    if not args:
        cmdline.print_help()
    else:
        for arg in args:
            tablist = tablistfile(arg, opt.inputformat)
            xlsx = arg + ".xlsx"
            workbook = tablistmake_workbook(tablist, opt.labels)
            if workbook:
                workbook.save(xlsx)
                rows = sum([len(tabtext.data) for tabtext in tablist])
                logg.log(DONE, " @ %s: %3d rows (%i tables)", xlsx, rows, len(tablist))
