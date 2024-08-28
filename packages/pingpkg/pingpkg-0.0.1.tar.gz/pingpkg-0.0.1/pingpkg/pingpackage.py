import os
import time
# Check if required libraries are installed, if not, install them
try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    os.system('pip install openpyxl')
    import openpyxl
    from openpyxl.styles import PatternFill

class pingToos():
    def __init__(self,ip="127.0.0.1",saveFilePath='C:\\Users\\Administrator\\Desktop\\',pingTime = 5,isTouchFile = 1)->None:
        self._ip = ip
        self._saveFilePath = saveFilePath
        self._pingTime = pingTime
        self._isTouchFile = isTouchFile
        
    def __touch_file(self) -> None:
        if os.path.exists(self._saveFilePath + 'ping_result.xlsx'):
            self.wb = openpyxl.load_workbook(self._saveFilePath + 'ping_result.xlsx')
            self.ws = self.wb.active
        else:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.append(['时间', '连接状态'])
    
    def __write_to_file(self,ping_result) -> None:
        if ping_result == 0:
            self.ws.append([time.strftime('%Y-%m-%d %H:%M:%S'), 'connect successful'])
        else:
            self.ws.append([time.strftime('%Y-%m-%d %H:%M:%S'), 'connect error'])
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            # Set red fill for columns A and B
            self.ws.cell(row=self.ws.max_row, column=1).style.fill = red_fill
            self.ws.cell(row=self.ws.max_row, column=2).style.fill = red_fill
        self.wb.save(self._saveFilePath + 'ping_result.xlsx')

    def __to_ping(self,save) -> int:
        response = os.system(f"ping -n 1 {self._ip}")
        if save:
            self.__write_to_file(response)
        else:
            return response

    def run(self) -> int:
        if self._isTouchFile:
            self.__touch_file()
            self.__to_ping(self._isTouchFile)
        else:
            return self.__to_ping(self._isTouchFile)
        
if __name__ == '__main__':
    testClass = pingToos(ip="example.com",saveFilePath="D:\\project\\pyproject\\windows_project\\get_time\\",isTouchFile=0)
    print(testClass.run())