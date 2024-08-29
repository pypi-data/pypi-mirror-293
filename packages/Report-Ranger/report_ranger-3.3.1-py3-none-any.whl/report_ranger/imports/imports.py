import os
from pathlib import Path
import logging
import csv
import traceback
import copy

try:
    from openpyxl import load_workbook
    openpyxl_imported = True
except ImportError:
    openpyxl_imported = False

log = logging.getLogger(__name__)


def import_csv(csvfile, as_dict_list=False, index_col=None):
    filepath = Path(csvfile)
    cwdpath = Path(os.path.curdir)
    if cwdpath not in filepath.parents:
        log.warning(
            "import_csv figure path {} doesn't seem to be in cwd".format(csvfile))
        return ''

    try:
        with open(filepath) as csvfileopen:
            reader = csv.reader(csvfileopen)
            return make_structure(reader, as_dict_list, index_col)
    except Exception as e:
        log.warning(
            "Error opening CSV file {}: {}".format(filepath, e.args))
        return ''


def import_xlsx(fp, worksheet=None, min_row=None, max_row=None, min_col=None, max_col=None, as_dict_list=False, index_col=None):
    if worksheet:
        log.info(f"Importing XLSX document {fp} worksheet {worksheet}")
    else:
        log.info(f"Importing XLSX document {fp}")
    if not openpyxl_imported:
        log.warning(
            "Unsuccessful loading of openpyxl, skipping import of XLSX file")
        return []
    filepath = Path(fp)
    cwdpath = Path(os.path.curdir)
    if cwdpath not in filepath.parents:
        log.warning(
            "import_xlsx path {} doesn't seem to be in cwd".format(fp))
        return ''

    try:
        workbook = load_workbook(filename=fp)
        if worksheet:
            sheet = workbook[worksheet]
        else:
            sheet = workbook.active

        row_iterator = sheet.iter_rows(
            min_row, max_row, min_col, max_col, values_only=True)
        return make_structure(row_iterator, as_dict_list, index_col)
    except Exception as e:
        log.warning(
            "error importing XLSX file {}: {}".format(fp, e.args))
        log.warn(traceback.format_exc())

        return ''


def make_structure(row_iterator, as_dict_list=False, index_col=None):
    # Let's just make 0 and 1 the same thing for simplicity and less user errors
    if index_col == 0:
        index_col = 1

    if index_col:
        results = {}
    else:
        results = []

    headings = []
    for row in row_iterator:
        # Make sure row is a list and not a tuple
        row = copy.deepcopy(list(row))

        # We need to store the headings to use as keys in our dicts.
        if as_dict_list and len(headings) < 1:
            headings = row
            continue

        # Turn this row into a dict, rather than a list.
        if as_dict_list:
            new_item = {}
            i = 0
            for key in headings:
                new_item[key] = row[i]
                i += 1
        else:
            new_item = row

        # We either create a new key/value pair or add the item to the list
        if index_col:
            results[row[index_col-1]] = new_item
        else:
            results.append(new_item)

    return results
