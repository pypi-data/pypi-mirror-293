import openpyxl


class ExcelReader:

    def read_excel_to_dict(self, file_path, sheet_name=None):

        workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

        if sheet_name:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        rows_as_dicts = []

        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

        for row in sheet.iter_rows(min_row=2):
            row_dict = {headers[i]: cell.value for i, cell in enumerate(row)}
            rows_as_dicts.append(row_dict)

        return rows_as_dicts
