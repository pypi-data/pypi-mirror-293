from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import logging
import json
import os
import datetime as dt
from statistics import mean
import decimal
from pathlib import Path
from binubuo import binubuo
from binubuo.BinubuoTemplate import BinubuoTemplate

class binubuoExcel(binubuo):
    def __init__(self, binubuokey):
        super().__init__(binubuokey)
        self.binubuokey = binubuokey
        self.max_sample_size = 100
        self.connected = False
        self.show_messages = True

    def setup_logging(self):
        # Create the logger
        self.logger = logging.getLogger(__name__)
        # Logger settings and levels
        self.logger.setLevel(logging.DEBUG)
        # Logger formatter and handler
        file_handler = logging.FileHandler(filename='binubuo.log', mode='a', encoding='utf8')
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)
        # Add the handler to the logger
        self.logger.addHandler(file_handler)

    def excel_line_count(self):
        return len([row for row in self.ws if not all([cell.value is None for cell in row])])

    def can_float(self, x):
        try:
            float(x)
            return True
        except:
            return False

    def is_date_string(self, str_in):
        date_fmts = ('%Y-%m-%dT%H:%M:%S.%f', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y %H:%M:%S', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y','%b %d, %Y','%b %d, %Y','%B %d, %Y','%B %d %Y','%m/%d/%Y','%m/%d/%y','%b %Y','%B%Y','%b %d,%Y')
        parsed = []
        if str_in is not None:
            for fmt in date_fmts:
                try:
                    t = dt.datetime.strptime(str(str_in), fmt)
                    parsed.append((str_in, fmt, t)) 
                    break
                except ValueError as err:
                    pass
        return parsed

    def templateFromFile(self, file_name, output_type="json"):
        self.m("Create template from Excel file [" + file_name + "].", tolog=True)
        table_template = BinubuoTemplate()
        # Open file and go through it.
        try:
            self.wb = load_workbook(filename = file_name)
            self.logger.info("File %s opened successfully", file_name)
            self.ws = self.wb.active
            # Row count
            row_count = self.excel_line_count()
            self.logger.info("File %s has %s rows with data.", file_name, row_count)
            # TODO: Build header check function here. For now header is not supported.
            self.current_has_header = False
            # Define store
            excel_sample_store = {}
            col_count = 0
            self.logger.debug("Start loading the sample store.")
            count = 0
            # We are looping column wise here
            for col in self.ws.iter_cols(min_row=1):
                col_count += 1
                self.logger.debug("setting up data store with generated headers")
                head_name = "c" + str(col_count)
                excel_sample_store[head_name] = []
                for cell in col:
                    count += 1
                    if count == int(self.binubuo_config.get('EXCEL', 'max_row_sample', fallback=100)):
                        break
                    else:
                        if not cell.value:
                            excel_sample_store[head_name].append(None)
                        else:
                            excel_sample_store[head_name].append(str(cell.value))
            # Now we have the sample store. Enumerate each to get infer data.
            self.logger.debug("Enumerate sample store for infer JSON generation")
            for count, idx in enumerate(excel_sample_store):
                self.logger.debug("Sample data for col %s: %s", idx, excel_sample_store[idx])
                infer_meta = {}
                self.logger.debug("Start creating infer JSON for column %s", idx)
                infer_meta["column_name"] = idx
                infer_meta["table_name"] = Path(file_name).stem
                # Nulls before checking for datatype
                self.logger.debug("before null check")
                infer_meta["column_nulls"] = sum(x is None for x in excel_sample_store[idx])
                if infer_meta["column_nulls"] > 0:
                    infer_meta["column_nullable"] = "Y"
                else:
                    infer_meta["column_nullable"] = "N"
                # Datatype check
                nums_chk = [x for x in excel_sample_store[idx] if self.can_float(x)]
                self.logger.debug("Numbers checked: %s", nums_chk)
                self.logger.debug("before datatype check")
                if len(nums_chk) == len(excel_sample_store[idx]) or len(nums_chk) == (len(excel_sample_store[idx]) - infer_meta["column_nulls"]):
                    infer_meta["column_datatype"] = "number"
                    infer_meta["column_number_decimals"] = round(mean([decimal.Decimal(str(x)).as_tuple().exponent for x in nums_chk])*-1)
                    self.logger.debug("Column %s is datatype: %s", idx, "number")
                else:
                    # TODO: Here we need to check if it is a date, before we just set it to string
                    # Checking for date format is very intensive. So as soon as we have one fail (len(0)) returned from the checker, we give up.
                    is_date = None
                    date_format = None
                    for str_num, str_val in enumerate(excel_sample_store[idx]):
                        self.logger.debug("About to date check %s", str_val)
                        has_date_format = self.is_date_string(str_val)
                        if len(has_date_format) == 0:
                            is_date = False
                            break
                        else:
                            is_date = True
                            date_format = has_date_format[0][1]
                            self.logger.debug("Found format: %s", date_format)
                    if is_date:
                        infer_meta["column_datatype"] = "date"
                        self.logger.debug("Column %s is datatype: %s", idx, "date")
                    else:
                        infer_meta["column_datatype"] = "string"
                        self.logger.debug("Column %s is datatype: %s", idx, "string")
                self.logger.debug("before data metadata settings")
                infer_meta["table_level_rowcount"] = row_count
                infer_meta["table_level_avg_row_length"] = ""
                infer_meta["column_level_avg_col_length"] = round(sum( map(len, map(str, excel_sample_store[idx])) ) / len(excel_sample_store[idx]))
                infer_meta["column_data_length"] = "4000"
                infer_meta["column_number_precision"] = ""
                # Unique list
                u_set = set(excel_sample_store[idx])
                u_list = (list(u_set))
                # Bug with unique. For now ignore
                #if len(u_list) == len(excel_sample_store[idx]):
                #    infer_meta["column_is_unique"] = "yes"
                #    infer_meta["column_distinct_count"] = len(excel_sample_store[idx])
                #else:
                infer_meta["column_is_unique"] = "no"
                infer_meta["column_distinct_count"] = len(excel_sample_store[idx]) - len(u_list)
                self.logger.debug("before data low/high settings")
                if infer_meta["column_datatype"] == "number":
                    infer_meta["column_low_value"] = min([float(x) for x in nums_chk])
                    infer_meta["column_high_value"] = max([float(x) for x in nums_chk])
                elif infer_meta["column_datatype"] == "date":
                    self.logger.debug("Just before date min/max")
                    infer_meta["column_low_value"] = min([dt.datetime.strptime(x, date_format) for x in excel_sample_store[idx] if x is not None]).strftime('%Y-%m-%d %H:%M:%S')
                    infer_meta["column_high_value"] = max([dt.datetime.strptime(x, date_format) for x in excel_sample_store[idx] if x is not None]).strftime('%Y-%m-%d %H:%M:%S')
                elif infer_meta["column_datatype"] == "string":
                    infer_meta["column_low_value"] = min([x for x in excel_sample_store[idx] if x is not None])
                    infer_meta["column_high_value"] = max([x for x in excel_sample_store[idx] if x is not None])
                if infer_meta["column_datatype"] == "number":
                    infer_meta["sample_values"] = nums_chk
                else:
                    infer_meta["sample_values"] = excel_sample_store[idx]
                self.logger.debug("Infer: %s", json.dumps(infer_meta))
                table_template.init_column(column_name=idx)
                x_response = self.infer_generator(json.dumps(infer_meta))
                self.logger.info("Infer result: %s", x_response)
                table_template.replace_column_from_json(idx, json.dumps(x_response))
            self.m("Template for excel file [" + file_name +"] successfully created.", indent=True, tolog=False)
            if output_type.upper() == "JSON":
                table_template.validate_template()
                table_template.complete_template()
                return table_template.template_JSON
            else:
                return table_template
        except Exception as e:
            self.logger.error("Error %s happened during read.", e)

    def dataset_from_excel(self, file_name):
        dset_name = Path(file_name).stem.lower()
        dset_template = self.templateFromFile(file_name)
        self.create_dataset(dset_name, dset_template)

    def dataset_to_excel(self, dataset_name, file_name, sheet_name=None, dataset_type="custom", dataset_category=None):
        self.m("Create new excel file [" + file_name + "] of Dataset  [" + dataset_name + "].", tolog=True)
        if (dataset_type.upper() == "CUSTOM" and self.dataset_exists(dataset_name)) or (self.standard_dataset_exists(dataset_name, dataset_category)):
            # Dataset exists and we can continue
            if not os.path.exists(file_name):
                # File does not exist, we can proceed.
                # Get the data
                excel_data = self.dataset(dataset_name = dataset_name, dataset_type = dataset_type, dataset_category = dataset_category)
                # create the workbook
                new_wb = Workbook()
                new_ws = new_wb.active
                if sheet_name is None:
                    new_ws.title = dataset_name
                else:
                    new_ws.title = sheet_name
                # Append the rows from the dataset
                for rw in excel_data:
                    new_ws.append(rw)
                # Now save the new file
                new_wb.save(file_name)
                self.m("New excel file [" + file_name + "] created successfully.", tolog=True)
            else:
                self.logger.error("File %s already exist. Please choose another filename or remove existing file first.", file_name)
        else:
            self.logger.error("Dataset requested %s does not exist.", dataset_name)

    def copy_csv(self, file_name, new_name="newsheet"
            , drop_target_if_exist=False, data_rows="source"):
        self.m("Create copy [" + new_name + "] of Excel file  [" + file_name + "].", tolog=True)
        dset_name = Path(file_name).stem
        can_create = False
        self.logger.debug("Before checking existence")
        if not self.dataset_exists(dset_name):
            self.dataset_from_excel(file_name)
        if new_name.upper() != "NEWSHEET" or new_name.upper() != "SHEET":
            # We are writing to a new file. Check if it exists.
            self.logger.debug("Check if %s already exists", new_name)
            if os.path.exists(new_name):
                if drop_target_if_exist:
                    # remove file
                    self.logger.debug("%s already exists, but allow dropped set. Removing file.", new_name)
                    os.remove(new_name)
                    can_create = True
            else:
                self.logger.debug("%s does not exists.", new_name)
                can_create = True
        else:
            # Using keyword location
            can_create = True
            self.logger.debug("Keyword location chosen: %s", new_name)
        if can_create:
            # Set the rows
            if data_rows == "source":
                self.drows(self.excel_line_count(file_name))
            else:
                self.drows(int(data_rows))
            # Get the data
            excel_data = self.dataset(dset_name)
            if new_name.upper() == "NEWSHEET":
                # Create a new sheet in the existing file.
                wb = load_workbook(filename = file_name)
                self.logger.info("File %s opened successfully", file_name)
                new_ws = wb.create_sheet(title=dset_name)
                # Append the rows from the dataset
                for rw in excel_data:
                    new_ws.append(rw)
                wb.save(file_name)
            elif new_name.upper() == "SHEET":
                # Append to existing sheet in existing file.
                wb = load_workbook(filename = file_name)
                self.logger.info("File %s opened successfully", file_name)
                ws = wb.active
                for rw in excel_data:
                    ws.append(rw)
                wb.save(file_name)
            else:
                # Create a new file
                new_wb = Workbook()
                new_ws = new_wb.active
                new_ws.title = dset_name
                # Append the rows from the dataset
                for rw in excel_data:
                    new_ws.append(rw)
                # Now save the new file
                new_wb.save(new_name)
            self.m("Copy [" + new_name + "] created successfully.", tolog=True)

