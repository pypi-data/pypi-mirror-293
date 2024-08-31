from docx import Document # python-docx
from io import BytesIO
from logyca_ai.utils.constants.ocr import OCREngine, OCREngineSettings
from openpyxl import load_workbook
from PIL import Image  # Pillow
import os
import pytesseract

def extract_text_from_docx_file(filename_full_path:str,advanced_image_recognition:bool=False,ocr_engine_path:str=None,output_temp_dir:str=None):
    """
    Extracts text from a DOCX file.

    :param filename_full_path: Full path to the DOCX file from which to extract text.
    :type filename_full_path: str
    :param advanced_image_recognition: Indicates whether to perform text recognition on images within the DOCX.
                               If True, OCR techniques will be used to extract text from images.
    :type advanced_image_recognition: bool
    :param ocr_engine_path: Path to the OCR executable. If provided, this path will be used instead of the default.
    :type ocr_engine_path: str, optional
    :param output_temp_dir: Temporary directory for storing output files.
                            If not provided, a default tmp temporary directory in the application root folder will be used.
    :type output_temp_dir: str, optional

    :return: Extracted text from the DOCX file.
    :rtype: str

    :raises FileNotFoundError: If the specified DOCX file is not found.
    :raises ValueError: If the OCR path is invalid.
    """

    if advanced_image_recognition:
        if ocr_engine_path is None:
            pytesseract.pytesseract.tesseract_cmd=OCREngine.get_binary_path()
        else:
            pytesseract.pytesseract.tesseract_cmd=ocr_engine_path

    doc = Document(filename_full_path)

    text = ""
    if output_temp_dir is None:
        output_temp_dir=os.path.abspath(os.path.join(os.path.dirname(__file__),OCREngineSettings.TMP_DIR))
    if not os.path.exists(output_temp_dir):
        os.makedirs(output_temp_dir)

    for paragraph in doc.paragraphs:
        text += paragraph.text + "\n"

    if advanced_image_recognition:
        for rel in doc.inline_shapes:
            if rel.type.value == 3:  # Type 3 corresponds to images
                rId = rel._inline.graphic.graphicData.pic.blipFill.blip.embed
                image_part = doc.part.related_parts[rId]
                image_bytes = image_part.image.blob

                image = Image.open(BytesIO(image_bytes))
                ocr_text = pytesseract.image_to_string(image)
                text += ocr_text

    return text

def extract_text_from_excel_file(filename_full_path: str, advanced_image_recognition: bool = False, ocr_engine_path: str = None, output_temp_dir: str = None):
    """
    Extracts text from an Excel file including all sheets and any embedded images.

    :param filename_full_path: Full path to the Excel file from which to extract text.
    :param advanced_image_recognition: Indicates whether to perform OCR on images within the Excel file.
    :param ocr_engine_path: Path to the OCR executable.
    :param output_temp_dir: Temporary directory for storing output files.
    :return: Extracted text from the Excel file.
    :rtype: str
    """
    # Set OCR engine path
    if advanced_image_recognition:
        if ocr_engine_path is None:
            pytesseract.pytesseract.tesseract_cmd = OCREngine.get_binary_path()
        else:
            pytesseract.pytesseract.tesseract_cmd = ocr_engine_path

    if output_temp_dir is None:
        output_temp_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), OCREngineSettings.TMP_DIR))
    if not os.path.exists(output_temp_dir):
        os.makedirs(output_temp_dir)

    text = ""

    workbook = load_workbook(filename_full_path, data_only=True)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        text += f"Sheet: {sheet_name}\n"
        
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    text += str(cell) + " "

        if advanced_image_recognition:
            for image in sheet._images:
                img = Image.open(BytesIO(image._data()))
                ocr_text = pytesseract.image_to_string(img)
                text += ocr_text

        text += "\n"

    return text